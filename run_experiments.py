#!/usr/bin/env python3
"""
run_experiments.py

Runs 3 Chaos Mesh experiments against a 3-node ZooKeeper StatefulSet.
Leader/follower roles are detected at runtime so the correct pod is always
targeted regardless of which node holds leadership.

Outputs
-------
  results.xlsx          -- multi-sheet Excel workbook (primary output)
  results.csv           -- flat summary row per experiment
  workload_<name>.log   -- raw workload.py output per experiment
"""

import csv
import queue
import re
import socket
import subprocess
import sys
import threading
import time
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# -- Configuration -------------------------------------------------------------

ZK_PODS   = ["zookeeper-0", "zookeeper-1", "zookeeper-2"]
ZK_BIN    = "/apache-zookeeper-3.9.3-bin/bin/zkServer.sh"
NAMESPACE = "default"

EXPERIMENTS = [
    {
        "name":          "kill_follower",
        "yaml":          "chaos_kill_follower.yaml",
        "kind":          "PodChaos",
        "resource_name": "kill-zk-follower",
        "chaos_type":    "pod_kill",
        "target_role":   "follower",
    },
    {
        "name":          "kill_leader",
        "yaml":          "chaos_kill_leader.yaml",
        "kind":          "PodChaos",
        "resource_name": "kill-zk-leader",
        "chaos_type":    "pod_kill",
        "target_role":   "leader",
    },
    {
        "name":          "network_partition",
        "yaml":          "chaos_network_partition.yaml",
        "kind":          "NetworkChaos",
        "resource_name": "zk-leader-partition",
        "chaos_type":    "network_partition",
        "target_role":   "leader",
    },
]

RESULTS_XLSX       = "results.xlsx"
RESULTS_CSV        = "results.csv"
RECOVERY_TIMEOUT   = 120   # seconds to wait for workload to recover
WORKLOAD_STABILIZE = 5     # seconds of baseline OK traffic before injecting chaos
POST_RECOVERY_BUF  = 5     # seconds to keep collecting after recovery
EXPERIMENT_GAP     = 20    # seconds between experiments

CSV_FIELDS = [
    "experiment_name", "description", "target_pod",
    "chaos_apply_time", "first_error_time", "recovery_time",
    "time_to_first_error_s", "recovery_duration_s", "total_outage_s",
    "pre_chaos_ok_count", "chaos_error_count", "chaos_ok_count",
    "unique_error_count",
    "pre_zk0", "pre_zk1", "pre_zk2",
    "post_zk0", "post_zk1", "post_zk2",
    "notes",
]

# Excel colour palette
CLR_HEADER     = "2F5496"   # dark blue
CLR_SUBHEADER  = "BDD7EE"   # light blue
CLR_OK         = "E2EFDA"   # light green
CLR_ERROR      = "FCE4D6"   # light salmon
CLR_INFO       = "F2F2F2"   # light grey
CLR_LEADER     = "FFD966"   # amber
CLR_FOLLOWER   = "DDEBF7"   # pale blue
CLR_DEAD       = "FF9999"   # light red
CLR_ALT        = "EBF3FB"   # alternate row tint


# -- Logging -------------------------------------------------------------------

run_start_dt = datetime.now()

def log(msg: str) -> None:
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


# -- ZooKeeper helpers ---------------------------------------------------------

def get_zk_status(pod: str) -> str:
    try:
        r = subprocess.run(
            ["kubectl", "exec", pod, "--", ZK_BIN, "status"],
            capture_output=True, text=True, timeout=15,
        )
        c = (r.stdout + r.stderr).lower()
        if "leader"      in c: return "leader"
        if "follower"    in c: return "follower"
        if "not running" in c: return "not_running"
        if "error"       in c: return "error"
        return c.strip()[:60] or "unknown"
    except subprocess.TimeoutExpired:
        return "timeout"
    except Exception as e:
        return f"error:{e}"


def get_all_statuses() -> dict:
    return {pod: get_zk_status(pod) for pod in ZK_PODS}


def resolve_roles(statuses: dict) -> tuple:
    leader    = next((p for p, s in statuses.items() if s == "leader"), None)
    followers = [p for p, s in statuses.items() if s == "follower"]
    return leader, followers


# -- Chaos helpers -------------------------------------------------------------

def build_pod_kill_yaml(resource_name: str, pod_name: str) -> str:
    return f"""\
apiVersion: chaos-mesh.org/v1alpha1
kind: PodChaos
metadata:
  name: {resource_name}
  namespace: {NAMESPACE}
spec:
  action: pod-kill
  mode: one
  gracePeriod: 0
  selector:
    namespaces:
      - {NAMESPACE}
    expressionSelectors:
      - key: statefulset.kubernetes.io/pod-name
        operator: In
        values:
          - {pod_name}
"""


def build_network_partition_yaml(resource_name: str,
                                  leader_pod: str,
                                  follower_pods: list) -> str:
    lines = "\n".join(f"          - {p}" for p in follower_pods)
    return f"""\
apiVersion: chaos-mesh.org/v1alpha1
kind: NetworkChaos
metadata:
  name: {resource_name}
  namespace: {NAMESPACE}
spec:
  action: partition
  mode: all
  selector:
    namespaces:
      - {NAMESPACE}
    expressionSelectors:
      - key: statefulset.kubernetes.io/pod-name
        operator: In
        values:
          - {leader_pod}
  direction: both
  target:
    mode: all
    selector:
      namespaces:
        - {NAMESPACE}
      expressionSelectors:
        - key: statefulset.kubernetes.io/pod-name
          operator: In
          values:
{lines}
  duration: "60s"
"""


def write_and_apply(yaml_content: str, yaml_path: str) -> None:
    with open(yaml_path, "w") as f:
        f.write(yaml_content)
    subprocess.run(["kubectl", "apply", "-f", yaml_path], check=True)


def delete_chaos(kind: str, name: str) -> None:
    subprocess.run(
        ["kubectl", "delete", kind, name, "--ignore-not-found=true"],
        check=True,
    )


# -- Workload subprocess -------------------------------------------------------

def start_workload(log_path: str):
    log_fh = open(log_path, "w")
    proc   = subprocess.Popen(
        [sys.executable, "-u", "workload.py"],
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True, bufsize=1,
    )
    q: queue.Queue = queue.Queue()

    def _reader():
        for raw in proc.stdout:
            line = raw.rstrip()
            log_fh.write(line + "\n")
            log_fh.flush()
            q.put(line)
        log_fh.close()
        q.put(None)

    threading.Thread(target=_reader, daemon=True).start()
    return proc, q


def stop_workload(proc) -> None:
    proc.terminate()
    try:
        proc.wait(timeout=5)
    except subprocess.TimeoutExpired:
        proc.kill()


# -- Event parsing -------------------------------------------------------------

def parse_workload_ts(line: str):
    try:
        ts_str = line[1:13]
        today  = datetime.now().strftime("%Y-%m-%d")
        return datetime.strptime(f"{today} {ts_str}", "%Y-%m-%d %H:%M:%S.%f")
    except Exception:
        return None


def classify_line(line: str):
    """Return (event_type, detail) for a workload output line."""
    ts = parse_workload_ts(line)
    if ts:
        if "] OK" in line:
            m = re.search(r"wrote (\d+)", line)
            return "OK", (f"counter={m.group(1)}" if m else "")
        if "] ERROR" in line:
            detail = line.split("] ERROR - ", 1)[-1] if "] ERROR - " in line else line
            return "ERROR", detail
    return "INFO", line


def make_event(exp_name, phase, raw_line, chaos_apply_dt):
    ts              = parse_workload_ts(raw_line)
    evt_type, detail = classify_line(raw_line)
    secs = ""
    if ts and chaos_apply_dt:
        v = round((ts - chaos_apply_dt).total_seconds(), 3)
        secs = v
    return {
        "experiment":        exp_name,
        "phase":             phase,
        "timestamp":         ts.strftime("%H:%M:%S.%f")[:-3] if ts else "",
        "type":              evt_type,
        "detail":            detail,
        "secs_since_chaos":  secs,
        "raw_line":          raw_line,
        "_ts":               ts,          # internal, not written to sheet
    }


def fmt_dt(dt) -> str:
    return dt.strftime("%H:%M:%S.%f")[:-3] if dt else "N/A"


def delta_s(a, b):
    if a and b:
        return round((b - a).total_seconds(), 3)
    return "N/A"


# -- Single experiment ---------------------------------------------------------

def run_experiment(exp: dict, all_events: list) -> dict | None:
    name = exp["name"]
    log(f"{'-'*60}")
    log(f"Experiment: {name}")
    log(f"{'-'*60}")

    delete_chaos(exp["kind"], exp["resource_name"])
    time.sleep(2)

    # Pre-chaos status + role detection
    log("Collecting pre-chaos ZK status ...")
    pre             = get_all_statuses()
    leader, followers = resolve_roles(pre)
    for pod, s in pre.items():
        log(f"  {pod}: {s}")

    if leader is None:
        log("ERROR: no leader detected — skipping.")
        return None

    log(f"  -> leader: {leader}  |  followers: {followers}")

    # Resolve target + build YAML
    if exp["target_role"] == "leader":
        target_pod = leader
    else:
        if not followers:
            log("ERROR: no followers detected — skipping.")
            return None
        target_pod = followers[0]

    if exp["chaos_type"] == "pod_kill":
        yaml_content = build_pod_kill_yaml(exp["resource_name"], target_pod)
        description  = f"Kill {target_pod} ({exp['target_role']})"
    else:
        others       = [p for p in ZK_PODS if p != target_pod]
        yaml_content = build_network_partition_yaml(
            exp["resource_name"], target_pod, others)
        description  = f"Network partition: isolate {target_pod} (leader) from {others}"

    log(f"Scenario: {description}")

    # Start workload
    log_path = f"workload_{name}.log"
    log(f"Starting workload.py -> {log_path}")
    proc, line_q = start_workload(log_path)

    # ── Stabilisation phase ────────────────────────────────────────────────
    exp_events   = []
    ok_seen      = False
    pre_ok_count = 0
    stab_end     = time.monotonic() + WORKLOAD_STABILIZE

    while time.monotonic() < stab_end:
        try:
            line = line_q.get(timeout=0.3)
        except queue.Empty:
            continue
        if line is None:
            log("ERROR: workload.py exited during stabilisation.")
            stop_workload(proc)
            return None
        log(f"  [workload] {line}")
        evt = make_event(name, "pre_chaos", line, None)
        exp_events.append(evt)
        if evt["type"] == "OK":
            ok_seen = True
            pre_ok_count += 1

    if not ok_seen:
        log("WARNING: no OK lines during stabilisation — ZooKeeper may be unhealthy.")

    # ── Inject chaos ───────────────────────────────────────────────────────
    log(f"Applying {exp['yaml']} ...")
    chaos_apply_dt = datetime.now()
    write_and_apply(yaml_content, exp["yaml"])
    log(f"Chaos applied at {fmt_dt(chaos_apply_dt)}")

    # ── Unified monitoring: collect all lines, track first ERROR + recovery ─
    first_error_dt = None
    recovery_dt    = None
    notes          = ""
    phase          = "chaos"        # transitions to "post_recovery" after OK
    deadline       = time.monotonic() + RECOVERY_TIMEOUT

    log("Monitoring workload ...")
    while time.monotonic() < deadline:
        try:
            line = line_q.get(timeout=0.3)
        except queue.Empty:
            continue
        if line is None:
            log("Workload process exited unexpectedly.")
            notes = "workload exited"
            break

        ts = parse_workload_ts(line)
        log(f"  [workload] {line}")
        evt = make_event(name, phase, line, chaos_apply_dt)
        exp_events.append(evt)

        if phase == "chaos":
            if evt["type"] == "ERROR" and first_error_dt is None:
                if ts and ts < chaos_apply_dt:
                    log("  (pre-chaos error, ignoring)")
                    exp_events[-1]["phase"] = "pre_chaos"
                else:
                    first_error_dt = ts or datetime.now()
                    log(f"  *** First disruption at {fmt_dt(first_error_dt)}")

            elif evt["type"] == "OK" and first_error_dt and recovery_dt is None:
                recovery_dt = ts or datetime.now()
                phase       = "post_recovery"
                deadline    = time.monotonic() + POST_RECOVERY_BUF
                log(f"  *** Recovery at {fmt_dt(recovery_dt)}")

        # post_recovery: keep collecting until deadline, then stop

    if first_error_dt is None and not notes:
        log("No disruption detected during chaos window.")
        notes = "no error detected"
    elif first_error_dt and recovery_dt is None and not notes:
        notes = f"recovery not observed within {RECOVERY_TIMEOUT}s"
        log(notes)

    # ── Post-recovery ZK status ────────────────────────────────────────────
    log("Collecting post-recovery ZK status ...")
    post = get_all_statuses()
    for pod, s in post.items():
        log(f"  {pod}: {s}")

    stop_workload(proc)
    delete_chaos(exp["kind"], exp["resource_name"])

    # ── Derive statistics from collected events ────────────────────────────
    chaos_evts      = [e for e in exp_events if e["phase"] == "chaos"]
    post_rec_evts   = [e for e in exp_events if e["phase"] == "post_recovery"]
    chaos_ok_count  = sum(1 for e in chaos_evts if e["type"] == "OK")
    chaos_err_count = sum(1 for e in chaos_evts if e["type"] == "ERROR")
    unique_errors   = list({e["detail"] for e in chaos_evts if e["type"] == "ERROR"})

    # Back-fill chaos_apply_dt into pre_chaos events (secs_since_chaos stays "")
    # and into chaos events (already set by make_event)

    all_events.extend(exp_events)

    result = {
        "experiment_name":       name,
        "description":           description,
        "target_pod":            target_pod,
        "chaos_apply_time":      fmt_dt(chaos_apply_dt),
        "first_error_time":      fmt_dt(first_error_dt),
        "recovery_time":         fmt_dt(recovery_dt),
        "time_to_first_error_s": delta_s(chaos_apply_dt, first_error_dt),
        "recovery_duration_s":   delta_s(first_error_dt, recovery_dt),
        "total_outage_s":        delta_s(chaos_apply_dt, recovery_dt),
        "pre_chaos_ok_count":    pre_ok_count,
        "chaos_error_count":     chaos_err_count,
        "chaos_ok_count":        chaos_ok_count,
        "unique_error_count":    len(unique_errors),
        "unique_errors":         unique_errors,          # not in CSV, used for Excel
        "pre_zk0":  pre["zookeeper-0"],
        "pre_zk1":  pre["zookeeper-1"],
        "pre_zk2":  pre["zookeeper-2"],
        "post_zk0": post["zookeeper-0"],
        "post_zk1": post["zookeeper-1"],
        "post_zk2": post["zookeeper-2"],
        "notes":    notes,
    }

    log(f"Outage: {result['total_outage_s']}s  "
        f"(to first error: {result['time_to_first_error_s']}s | "
        f"recovery: {result['recovery_duration_s']}s) | "
        f"errors: {chaos_err_count}  unique: {len(unique_errors)}")
    return result


# -- Excel generation ----------------------------------------------------------

def _hfill(hex_str: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_str)


def _header_font() -> Font:
    return Font(bold=True, color="FFFFFF")


def _bold() -> Font:
    return Font(bold=True)


def _set_col_widths(ws, widths: list) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_header_row(ws, headers: list, fill_hex: str = CLR_HEADER) -> None:
    fill = _hfill(fill_hex)
    font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30


def _status_fill(status: str) -> PatternFill | None:
    s = (status or "").lower()
    if s == "leader":     return _hfill(CLR_LEADER)
    if s == "follower":   return _hfill(CLR_FOLLOWER)
    if "not_running" in s or "error" in s or "timeout" in s:
        return _hfill(CLR_DEAD)
    return None


def generate_excel(results: list, all_events: list, path: str) -> None:
    wb = Workbook()

    # ── Sheet 1: Run Info ──────────────────────────────────────────────────
    ws_info = wb.active
    ws_info.title = "Run Info"
    run_end_dt = datetime.now()
    rows = [
        ("Run date",            run_start_dt.strftime("%Y-%m-%d")),
        ("Run start",           run_start_dt.strftime("%H:%M:%S")),
        ("Run end",             run_end_dt.strftime("%H:%M:%S")),
        ("Total duration (s)",  round((run_end_dt - run_start_dt).total_seconds(), 1)),
        ("Experiments run",     len(results)),
        ("ZooKeeper pods",      ", ".join(ZK_PODS)),
        ("Recovery timeout (s)", RECOVERY_TIMEOUT),
        ("Stabilise window (s)", WORKLOAD_STABILIZE),
        ("Post-recovery buf (s)", POST_RECOVERY_BUF),
        ("Results file",        path),
    ]
    ws_info.column_dimensions["A"].width = 26
    ws_info.column_dimensions["B"].width = 40
    for r, (k, v) in enumerate(rows, start=1):
        ws_info.cell(r, 1, k).font = _bold()
        ws_info.cell(r, 2, v)

    # ── Sheet 2: Summary ───────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    sum_headers = [
        "Experiment", "Scenario", "Target Pod",
        "Chaos Applied", "First Error", "Recovery",
        "Time to Error (s)", "Recovery Duration (s)", "Total Outage (s)",
        "Pre-Chaos OKs", "Chaos Errors", "Chaos OKs", "Unique Error Types",
        "Pre ZK-0", "Pre ZK-1", "Pre ZK-2",
        "Post ZK-0", "Post ZK-1", "Post ZK-2",
        "Notes",
    ]
    _write_header_row(ws_sum, sum_headers)
    _set_col_widths(ws_sum, [
        18, 52, 14,
        12, 12, 12,
        16, 20, 16,
        14, 13, 10, 18,
        12, 12, 12,
        12, 12, 12,
        35,
    ])

    for row_i, r in enumerate(results, start=2):
        fill = _hfill(CLR_ALT) if row_i % 2 == 0 else None
        vals = [
            r["experiment_name"], r["description"], r["target_pod"],
            r["chaos_apply_time"], r["first_error_time"], r["recovery_time"],
            r["time_to_first_error_s"], r["recovery_duration_s"], r["total_outage_s"],
            r["pre_chaos_ok_count"], r["chaos_error_count"], r["chaos_ok_count"],
            r["unique_error_count"],
            r["pre_zk0"], r["pre_zk1"], r["pre_zk2"],
            r["post_zk0"], r["post_zk1"], r["post_zk2"],
            r["notes"],
        ]
        for col_i, v in enumerate(vals, start=1):
            cell = ws_sum.cell(row=row_i, column=col_i, value=v)
            if fill:
                cell.fill = fill
        # Colour ZK status cells (columns 14-19)
        for col_i, status_key in enumerate(
            ["pre_zk0","pre_zk1","pre_zk2","post_zk0","post_zk1","post_zk2"],
            start=14
        ):
            sf = _status_fill(r[status_key])
            if sf:
                ws_sum.cell(row=row_i, column=col_i).fill = sf
        # Highlight long outages (> 30s) in outage column (col 9)
        total = r["total_outage_s"]
        if isinstance(total, (int, float)) and total > 30:
            ws_sum.cell(row=row_i, column=9).fill = _hfill(CLR_ERROR)
            ws_sum.cell(row=row_i, column=9).font = _bold()

    # ── Sheet 3: Workload Events ───────────────────────────────────────────
    ws_ev = wb.create_sheet("Workload Events")
    ev_headers = [
        "Experiment", "Phase", "Timestamp", "Event Type",
        "Detail / Message", "Secs Since Chaos Applied",
    ]
    _write_header_row(ws_ev, ev_headers)
    _set_col_widths(ws_ev, [18, 16, 14, 12, 80, 22])

    type_fill = {"OK": _hfill(CLR_OK), "ERROR": _hfill(CLR_ERROR), "INFO": _hfill(CLR_INFO)}

    for row_i, e in enumerate(all_events, start=2):
        vals = [
            e["experiment"], e["phase"], e["timestamp"],
            e["type"], e["detail"], e["secs_since_chaos"],
        ]
        fill = type_fill.get(e["type"])
        for col_i, v in enumerate(vals, start=1):
            cell = ws_ev.cell(row=row_i, column=col_i, value=v)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(wrap_text=False)

    # ── Sheet 4: ZK Status History ─────────────────────────────────────────
    ws_zk = wb.create_sheet("ZK Status History")
    zk_headers = [
        "Experiment", "Phase",
        "zookeeper-0", "zookeeper-1", "zookeeper-2",
    ]
    _write_header_row(ws_zk, zk_headers)
    _set_col_widths(ws_zk, [20, 8, 16, 16, 16])

    for r in results:
        for phase, keys in [("pre",  ["pre_zk0",  "pre_zk1",  "pre_zk2"]),
                             ("post", ["post_zk0", "post_zk1", "post_zk2"])]:
            next_row = ws_zk.max_row + 1
            ws_zk.cell(next_row, 1, r["experiment_name"])
            ws_zk.cell(next_row, 2, phase)
            for col_i, key in enumerate(keys, start=3):
                cell = ws_zk.cell(next_row, col_i, r[key])
                sf = _status_fill(r[key])
                if sf:
                    cell.fill = sf

    # ── Sheet 5: Error Analysis ────────────────────────────────────────────
    ws_err = wb.create_sheet("Error Analysis")
    err_headers = [
        "Experiment", "Target Pod", "Total Outage (s)",
        "Chaos Errors", "Unique Error Types", "Error Message(s)",
        "Pre-Chaos OKs", "Chaos OKs",
    ]
    _write_header_row(ws_err, err_headers)
    _set_col_widths(ws_err, [18, 14, 16, 14, 18, 80, 14, 10])

    for row_i, r in enumerate(results, start=2):
        error_msgs = " | ".join(r.get("unique_errors", [])) or "none"
        vals = [
            r["experiment_name"], r["target_pod"], r["total_outage_s"],
            r["chaos_error_count"], r["unique_error_count"], error_msgs,
            r["pre_chaos_ok_count"], r["chaos_ok_count"],
        ]
        fill = _hfill(CLR_ALT) if row_i % 2 == 0 else None
        for col_i, v in enumerate(vals, start=1):
            cell = ws_err.cell(row=row_i, column=col_i, value=v)
            if fill:
                cell.fill = fill
        if r["chaos_error_count"] > 0:
            ws_err.cell(row=row_i, column=4).fill = _hfill(CLR_ERROR)

    wb.save(path)
    log(f"Excel workbook saved -> {path}  "
        f"({len(all_events)} events across {len(results)} experiments)")


# -- Main ----------------------------------------------------------------------

def check_port_forward() -> bool:
    try:
        s = socket.create_connection(("127.0.0.1", 2181), timeout=3)
        s.close()
        return True
    except OSError:
        return False


def main():
    if not check_port_forward():
        print(
            "\nERROR: Nothing is listening on 127.0.0.1:2181.\n"
            "Run this in a separate terminal first:\n\n"
            "    kubectl port-forward svc/zookeeper 2181:2181\n"
        )
        sys.exit(1)

    log("Port-forward on :2181 confirmed.")
    log(f"Outputs: {RESULTS_XLSX}  |  {RESULTS_CSV}  |  workload_<name>.log")
    log(f"Running {len(EXPERIMENTS)} experiments ...\n")

    with open(RESULTS_CSV, "w", newline="") as f:
        csv.DictWriter(f, fieldnames=CSV_FIELDS).writeheader()

    results: list   = []
    all_events: list = []

    for i, exp in enumerate(EXPERIMENTS):
        if i > 0:
            log(f"\nWaiting {EXPERIMENT_GAP}s for ZooKeeper to stabilise ...\n")
            time.sleep(EXPERIMENT_GAP)

        result = run_experiment(exp, all_events)
        if result is None:
            log(f"Experiment '{exp['name']}' skipped.")
            continue

        results.append(result)

        # Write CSV row (exclude Excel-only fields)
        csv_row = {k: result[k] for k in CSV_FIELDS}
        with open(RESULTS_CSV, "a", newline="") as f:
            csv.DictWriter(f, fieldnames=CSV_FIELDS).writerow(csv_row)

        log(f"Row appended to {RESULTS_CSV}.\n")

    log(f"{'-'*60}")
    generate_excel(results, all_events, RESULTS_XLSX)
    log(f"All experiments complete.")
    log(f"  {RESULTS_XLSX} -- {len(results)} experiments, {len(all_events)} events")
    log(f"  {RESULTS_CSV}  -- flat summary")


if __name__ == "__main__":
    main()
